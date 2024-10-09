import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from Validaciones import guarda_excel

# Rutas
nombre_archivo = "Errores.xlsx"
path_carpeta = r'C:\Users\magod\OneDrive\Escritorio\Museo\Scripts'
path_destino = path_carpeta + f'\\{nombre_archivo}'

# Color constants
RED_FILL = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
YELLOW_FILL = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
ORANGE_FILL = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')
GREEN_FILL = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')

def Reportes_F2(df):
    try:
        # Guardar el DataFrame a un archivo Excel
        df.to_excel(path_destino, index=False)

        # Cargar el archivo Excel con openpyxl
        wb = load_workbook(path_destino)
        ws = wb.active

        # Detectar las columnas "SpectrumID" y "ClasificacionID"
        spectrum_col = None
        clasificacion_col = None
        for col in ws.iter_cols(1, ws.max_column):
            if col[0].value == "ID SPECTRUM":
                spectrum_col = col[0].column
            elif col[0].value == "ID Clasificacion":
                clasificacion_col = col[0].column

        # Asegurarse de que ambas columnas se detectaron
        if spectrum_col is not None and clasificacion_col is not None:
            # Iterar sobre cada fila, comenzando desde la fila 2 para evitar los encabezados
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                spectrum_value = row[spectrum_col - 1].value
                clasificacion_value = row[clasificacion_col - 1].value

                # Definir el color de fondo según las condiciones
                if spectrum_value == 'otras' and clasificacion_value == 'error':
                    fill = RED_FILL
                elif spectrum_value == 'otras':
                    fill = YELLOW_FILL
                elif clasificacion_value in ['error0', 'error1', 'error2', 'error']:
                    fill = ORANGE_FILL
                else:
                    fill = GREEN_FILL

                # Aplicar el color de fondo
                for cell in row:
                    cell.fill = fill

        # Guardar el archivo Excel modificado
        wb.save(path_destino)
    except PermissionError:
        print(f"El archivo Excel {nombre_archivo} está abierto. Porfavor cierralo e intenta nuevamente.")
    except Exception as e:
        print(f"Un error inesperado ocurrió: {e}")